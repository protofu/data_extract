import csv
import io
import os
import yaml
from datetime import datetime, timedelta
import pymysql
import pandas as pd
import numpy as np
import re
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
import logging
# logging.basicConfig(level=logging.DEBUG)

from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError


class CollectionVolumeMonitor:

    @staticmethod
    def amount_of_change(values, alert):
        """변화량 계산"""
        # 기준일 수집량
        today_col = values[0]
        rlt = {}

        # 1일 대비 변화량 계산
        day_1 = today_col - values[1]
        day_1_per = '-' if values[1] == 0 else np.round((day_1 * 100) / values[1], 1)
        alert['day_delta'] = day_1
        alert['day_per'] = day_1_per

        # 평균 기준 계산
        for d in list(values.keys())[2:]:
            compari = values[d][1]
            change = today_col - compari
            if compari == 0:
                alert[f'{d}_day_delta'] = 0
                alert[f'{d}_day_per'] = 0
                alert[f'{d}_avg'] = 0
            else:
                percent = np.round((change * 100) / compari, 1)
                rlt.setdefault(d, (change, percent))
                alert[f'{d}_day_delta'] = change
                alert[f'{d}_day_per'] = percent
                alert[f'{d}_avg'] = compari
        return alert

    @staticmethod
    def avg_calc(today, next_day):
        amount_of_change = today - next_day
        if next_day == 0:
            return f"{amount_of_change}({amount_of_change * 100}%)"
        return f"{amount_of_change}({np.round(amount_of_change * 100 / next_day, 1)}%)"

    def __init__(self, config_f):
        if not os.path.exists(config_f):
            raise IOError(f'Cannot read config file "{config_f}"')
        with open(config_f, encoding='utf-8') as ifp:
            self.config = yaml.load(ifp, yaml.SafeLoader)

        date_time = datetime.today()
        today = date_time.date()

        # 슬랙 변수 관련 설정
        slack_params = self.config.get('params', {}).get('slack', {})
        self.slack_oauth_token = slack_params.get('OAuth_token', '')
        self.slack_mention_ids = slack_params.get('user_id', [])
        self.slack_channel = slack_params.get('channel', '')

        # DB 관련 변수 설정
        db_params = self.config.get('database', {})
        self.db_host = db_params.get('host', '')
        self.db_user = db_params.get('user', '')
        self.db_password = db_params.get('password', '')
        self.db_database = db_params.get('database', 'vocdailystats')

        # 분석 관련 변수 설정
        date_params = self.config.get('params',  {}).get('date', {})
        self.standard_date = date_params.get('standard_date', today)
        self.limit_day = date_params.get('limit', 60)   # 값이 없으면 기본 60 기준
        self.data_call_end_date = self.standard_date - timedelta(days=self.limit_day)
        self.window_sizes = date_params.get('window_sizes', [0, 1, 3, 7, 14, 30, 60])

        # 알람 조건 관련 설정
        alert_params = self.config.get('params', {}).get('alert', {})
        self.threshold = alert_params.get('threshold', 10)
        self.min_collect = alert_params.get('min_collect', 20)
        self.min_7_avg = alert_params.get('min_7_avg', 100)
        self.critical = alert_params.get('critical', {})
        self.warning = alert_params.get('warning', {})
        self.caution = alert_params.get('caution', {})

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        # 필요한 정리 작업 (예: DB 연결 종료)
        if hasattr(self, 'conn'):
            self.conn.close()
        if hasattr(self, 'cursor'):
            self.cursor.close()

    def extract_data(self):
        """MySQL로 부터 데이터 추출"""
        try:
            # MySQL 연결
            conn = pymysql.connect(
                host=self.db_host,
                user=self.db_user,
                password=self.db_password,
                database=self.db_database,
                charset='utf8'
            )

            # 커서 생성
            cursor = conn.cursor()
            # 데이터 추출 쿼리
            excute_query = f'''
                SELECT
                    site_number_and_name, 
                    collection_end_date,
                    COUNT(*) AS total_count
                FROM 
                    merge
                WHERE
                    collection_end_date BETWEEN %s AND %s 
                GROUP BY
                    site_number_and_name,
                    collection_end_date
                ORDER BY
                    collection_end_date DESC;
            '''
            cursor.execute(excute_query, (
                self.data_call_end_date,
                self.standard_date,
            ))

            rows = cursor.fetchall()
            cursor.close()
            conn.close()

            return rows

        except Exception as e:
            print(f"DB 연결 실패: {e}")
        finally:
            # 커서 종료
            if hasattr(self, 'cursor'):
                self.cursor.close()
            # 연결 종료
            if hasattr(self, 'conn'):
                self.conn.close()
                print("DB 연결 종료!")

    def data_preprocessing(self, data):
        """추출된 데이터 전처리"""
        prepro_data = {}
        start_date = self.standard_date
        end_date = self.data_call_end_date
        for site, date, collect in data:
            prepro_data.setdefault(site, []).append((date, collect))
        for key in prepro_data:
            date_range = [start_date - timedelta(days=i) for i in range((start_date - end_date).days + 1)]
            date_dict = {date: 0 for date in date_range}  # 기본값은 0
            for date, value in prepro_data[key]:
                if date in date_dict:
                    date_dict[date] = value
            result = [(date, date_dict[date]) for date in date_range]
            prepro_data[key] = result
        return prepro_data

    def calc_mid_avg_std(self, data):
        """평균, 중간, 표준편차 구하기"""
        mid_avg_std = {}
        for site, values in data.items():
            # 날짜와 값 추출
            collect = [row[1] for row in values]

            # numpy 배열로 변환
            collect_np = np.array(collect)

            # 이동평균 범위 변수값
            window_sizes = self.window_sizes

            # 계산
            for window_size in window_sizes:
                # 0일은 당일 수집량 담기
                if window_size == 0:
                    mid_avg_std.setdefault(site, {}).setdefault(window_size, collect_np[0])
                # 1일은 전날 수집량을 담기
                elif window_size == 1:
                    mid_avg_std.setdefault(site, {}).setdefault(window_size, collect_np[1])
                else:
                    # 기간을 넘지 않게 slice 하여 평균 계산
                    end_idx = window_size + 1
                    window_data = collect_np[0:end_idx]

                    m_val = np.median(window_data)
                    a_val = np.average(window_data)
                    s_val = np.std(window_data)
                    # 2개의 값이 다 NaN이 아니면 기록
                    if not np.isnan(a_val) and not np.isnan(m_val) and not np.isnan(s_val):
                        mid_avg_std.setdefault(site, {}).setdefault(window_size, []).append(int(np.round(m_val)))
                        mid_avg_std.setdefault(site, {}).setdefault(window_size, []).append(int(np.round(a_val)))
                        mid_avg_std.setdefault(site, {}).setdefault(window_size, []).append(int(np.round(s_val, 2)))
                    else:
                        mid_avg_std.setdefault(site, {}).setdefault(window_size, -1)

        return mid_avg_std

    def is_alert(self, data):
        """알람 데이터 판별"""
        alert_list = {
            'critical': [],
            'warning': [],
            'caution': [],
        }
        for site, val in data.items():
            alert = {
                'site_name': site,
                'today_collect': val[0],
                'footer': None
            }
            # 변화량 계산
            com_alert = self.amount_of_change(val, alert)
            # 알람 조건에 따라 거르기
            rlt = self.cond_of_alert(com_alert)
            if rlt is not None:
                alert_list[rlt['stage']].append(rlt)
        return alert_list

    def cond_of_alert(self, alert):
        """알람 조건"""
        # 7일 평균이 설정값 미만이라면 None 반환(제외).
        if alert['7_avg'] < self.min_7_avg:
            return None

        day_1 = alert['day_per'] if alert['day_per'] != '-' else 0
        day_3 = alert['3_day_per'] if alert['3_day_per'] != '-' else 0
        day_7 = alert['7_day_per'] if alert['7_day_per'] != '-' else 0
        alert['footer'] = ''

        # 단계별 조건과 메시지를 처리하는 함수
        def check_stage(stage, thresholds):
            conditions = [
                (abs(day_1) >= thresholds['day'], "전일,\t"),
                (abs(day_3) >= thresholds['3_day'], "3일,\t"),
                (abs(day_7) >= thresholds['7_day'], "7일")
            ]
            matched = False
            for condition, message in conditions:
                if condition:
                    nonlocal alert
                    alert['footer'] += message.format(day_1=day_1, day_3=day_3, day_7=day_7)
                    matched = True
            return matched

        # 1차 심각 단계 우선
        if check_stage('critical', self.critical):
            alert['stage'] = 'critical'
        # 2차 경계 단계
        elif check_stage('warning', self.warning):
            alert['stage'] = 'warning'
        # 3차 주의 단계
        elif check_stage('caution', self.caution):
            alert['stage'] = 'caution'
        else:
            return None

        return alert

    def csv_slack_msg(self, alert_list):
        """slack 메세지 전송"""
        # OAuth 토큰 설정
        slack_token = self.slack_oauth_token
        client = WebClient(token=slack_token)

        # 멘션할 사용자 ID
        mention_user_ids = ''
        for user in self.slack_mention_ids:
            mention_user_ids += f'<@{user}> '

        # Generate CSV data
        csv_buffer = io.BytesIO()  # BytesIO 사용
        writer = csv.writer(io.TextIOWrapper(csv_buffer, encoding='utf-8', write_through=True))
        csv_buffer.write(b'\xef\xbb\xbf')

        writer.writerow(["단계", "사이트명", "당일", "전일 대비", "3일 대비", "7일 대비", '특이 변화'])

        for stage, values in alert_list.items():
            for val in values:
                site_name = val['site_name']
                # 당일 수집량
                today = val['today_collect']

                # 값이 없으면 '--' 표시하고, 퍼센트는 우측 정렬을 위해 길이를 맞춤
                day_1 = f"{val.get('day_delta', '')}\t\t({val.get('day_per', '--')} %)" if val.get('day_per', '') else "--"
                day_3 = f"{val.get('3_day_delta', '')}\t\t({val.get('3_day_per', '--')} %)" if val.get('3_day_per', '') else "--"
                day_7 = f"{val.get('7_day_delta', '')}\t\t({val.get('7_day_per', '--')} %)" if val.get('7_day_per', '') else "--"


                if val.get('stage', '') == 'critical':
                    stage = '[심각]'
                elif val.get('stage', '') == 'warning':
                    stage = '[경계]'
                elif val.get('stage', '') == 'caution':
                    stage = '[주의]'
                footer = val['footer']

                writer.writerow([stage, site_name, today, day_1, day_3, day_7, footer])

        csv_buffer.seek(0)

        # '년 월 일' 형식으로 날짜 포매팅
        formatted_date = self.standard_date.strftime("%Y년 %m월 %d일")

        # CSV to Slack_msg
        try:
            response = client.files_upload_v2(
                channel=self.slack_channel,
                file=csv_buffer,
                filename="alert_data.csv",
                title="사이트 변화 감지 데이터",
                initial_comment=f"{mention_user_ids}\n{formatted_date} 사이트별 모니터링"
            )
        except SlackApiError as e:
            assert e.response["error"]

    def excel_saver(self, data):
        """엑셀 저장 함수"""
        after_data = {}
        for key, val in data.items():
            after_data.setdefault(key, {}).setdefault('당일', val[0])
            after_data[key]['전일'] = val[1]
            after_data[key]['7일'] = val[7][1]
            after_data[key]['14일'] = val[14][1]
            after_data[key]['30일'] = val[30][1]
            after_data[key]['60일'] = val[60][1]
            after_data[key]['7일 대비 변화량'] = self.avg_calc(val[0], val[7][1])

        df = pd.DataFrame(after_data).T

        # 엑셀 저장
        with pd.ExcelWriter(f'{self.standard_date} 기준 모니터링 결과.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='수집 데이터')

            # openpyxl로 엑셀 파일 접근
            # workbook = writer.book
            worksheet = writer.sheets['수집 데이터']

            # 스타일 정의
            number_style = NamedStyle(name="number_style", number_format='#,##0')  # 숫자 형식 (천 단위 구분)
            date_style = NamedStyle(name="date_style", number_format='YYYY-MM-DD')  # 날짜 형식
            change_style = NamedStyle(name="change_style", number_format='#,##0.0%')  # 변화량 퍼센트 형식
            bold_font = Font(bold=True)  # Bold 글꼴 스타일
            align_center = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
            align_right = Alignment(horizontal='right')
            align_left = Alignment(horizontal='left')

            # 음수 값에 파란색 배경 색상 적용
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            blue_font = Font(color="0000FF")
            red_font = Font(color="FF0000")

            # 헤더 스타일 적용 (볼드 처리, 중앙 정렬)
            for cell in worksheet[1]:
                cell.font = bold_font
                cell.alignment = align_center

            # 스타일 적용: 각 열에 맞는 스타일 적용
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    # 날짜 열 스타일 적용
                    if cell.column == 1:  # 첫 번째 열 (날짜 열)
                        cell.style = date_style
                        cell.font = bold_font
                        cell.alignment = align_left
                    # 변화량 열 스타일 적용
                    elif '변화량' in worksheet.cell(row=1, column=cell.column).value:  # 변화량 열
                        cell.style = change_style
                        cell.alignment = align_right
                        number = int(re.sub(r"\(.*\)", "", cell.value).strip())
                        # print(number, type(number))
                        if number < 0:
                            cell.font = blue_font
                        elif number > 0:
                            cell.font = red_font

                    else:
                        cell.style = number_style
                        cell.alignment = align_right

            # 1행의 헤더 열 폭을 설정
            for col in worksheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=worksheet.max_column):
                for cell in col:
                    # 각 열의 최대 길이를 계산
                    max_length = 0
                    column = cell.column_letter
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                        for c in row:
                            if c.column == cell.column:  # 같은 열의 셀들에 대해 최대 길이 계산
                                try:
                                    if len(str(c.value)) > max_length:
                                        max_length = len(str(c.value))
                                except:
                                    pass
                    adjusted_width = (max_length + 2)  # 여유 공간을 2만큼 더 추가하여 너비 설정
                    worksheet.column_dimensions[column].width = adjusted_width

    def start(self):
        db_datas = self.extract_data()
        if len(db_datas):
            prepro_datas = self.data_preprocessing(db_datas)
            # 중간, 평균, 표준편차 데이터 추출
            mas_datas = self.calc_mid_avg_std(prepro_datas)
            self.excel_saver(mas_datas)
            alert_list = self.is_alert(mas_datas)
            self.csv_slack_msg(alert_list)


def start_monitoring(**kwargs):
    with CollectionVolumeMonitor(kwargs['config_f']) as cvm:
        cvm.start()
        return 0


if __name__ == '__main__':
    _config_f = 'collection_volume_monitor.yaml'
    start_monitoring(config_f=_config_f)



